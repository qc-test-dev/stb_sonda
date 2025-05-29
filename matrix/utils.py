import openpyxl
from .models import CasoDePrueba,Validate
from requests.auth import HTTPBasicAuth
from decouple import config
import pandas as pd
import re
import requests
import random

def limpiar(valor):
    if isinstance(valor, str):
        return valor.strip()
    return valor

import openpyxl
from .models import CasoDePrueba

import openpyxl
from .models import CasoDePrueba

def importar_matriz_desde_excel(matriz, ruta_excel, alcances_permitidos=None):
    """
    Importa casos de prueba desde un archivo Excel y los asigna a una matriz.
    Filtra por alcance si se proporciona una lista de alcances_permitidos (['A', 'B', 'C']).
    Las filas incompletas (sin alcance, fase, caso o criticidad) se ignoran.
    """
    wb = openpyxl.load_workbook(ruta_excel)
    sheet = wb.active

    for fila in sheet.iter_rows(min_row=2, values_only=True):
        alcance = fila[0]
        fase = fila[1]
        caso_de_prueba = fila[2]
        criticidad = fila[4]
        nota = fila[5] if len(fila) > 5 else ""

        # Validar que los campos clave no estén vacíos
        if not (alcance and fase and caso_de_prueba and criticidad):
            continue  # Ignorar la fila si falta alguno

        # Filtrar por alcance si se especifica
        if alcances_permitidos and alcance not in alcances_permitidos:
            continue

        # Crear el caso de prueba
        CasoDePrueba.objects.create(
            matriz=matriz,
            alcance=alcance,
            fase=fase,
            caso_de_prueba=caso_de_prueba,
            estado="por_ejecutar",
            criticidad=criticidad,
            nota=nota or ""
        )


# def importar_validates_desde_excel(super_matriz, ruta_excel):
#     """
#     Importa registros de 'Validate' desde un archivo Excel y los asigna a una SuperMatriz.
#     Las filas incompletas (sin tester, ticket, descripcion, prioridad o estado) se ignoran.
#     """
#     wb = openpyxl.load_workbook(ruta_excel)
#     sheet = wb.active

#     empty_rows = 0

#     for fila in sheet.iter_rows(min_row=2, values_only=True):
#         if all(cell is None for cell in fila):
#             empty_rows += 1
#             if empty_rows > 5:
#                 break
#             #print("Fila ignorada por estar incompleta")
#             continue
#         empty_rows = 0  # Reset counter si hay datos

#         tester = fila[0]
#         ticket = fila[1]
#         descripcion = fila[2]
#         prioridad = fila[3]
#         estado = fila[4]

#         # Ignorar si falta alguno de los campos obligatorios
#         if not tester or not ticket or not descripcion or not prioridad or not estado:
#             #print("Fila ignorada por campos vacíos obligatorios")
#             continue

#         Validate.objects.create(
#             super_matriz=super_matriz,
#             tester=tester,
#             ticket=ticket,
#             descripcion=descripcion,
#             prioridad=prioridad,
#             estado=estado
#         )

def importar_validates(super_matriz, link):
    TESTERS = ['Kevin', 'Erik', 'Kyle', 'Alberto', 'Axel', 'Luis Rene']

    issues, error = fetch_jira_issues(link)

    if error:
        print(f"Error al obtener issues: {error}")
        return
    
    if not issues:
        print("No se encontraron issues.")
        return

    random.shuffle(issues)  # Baraja los issues, no necesitas barajar los testers

    tester_count = len(TESTERS)
    assignments = {tester: [] for tester in TESTERS}

    # Distribuye equitativamente
    for idx, caso in enumerate(issues):
        assigned_tester = TESTERS[idx % tester_count]
        assignments[assigned_tester].append(caso)

    # Crea los Validate en bloques por tester (mejor para la DB que hacer cada uno por separado)
    validate_objects = []
    for tester, casos in assignments.items():
        for caso in casos:
            validate_objects.append(
                Validate(
                    super_matriz=super_matriz,
                    tester=tester,
                    ticket=caso["key"],
                    descripcion=caso["summary"],
                    prioridad=caso['priority'],
                    estado="por_ejecutar"
                )
            )
    
    Validate.objects.bulk_create(validate_objects)
    print(f"Se crearon {len(validate_objects)} validates distribuidos entre testers.")

def fetch_jira_issues(link):
    """
    Obtiene los issues desde Jira usando un filtro.
    """
    match = re.search(r'filter=(\d+)', link)
    if not match:
        return None, "No se pudo extraer el filtro del enlace."

    filter_id = match.group(1)

    url = "https://dlatvarg.atlassian.net/rest/api/3/search"
    params = {'jql': f'filter = {filter_id}'}
    headers = {"Accept": "application/json"}

    try:
        response = requests.get(
            url,
            headers=headers,
            params=params,
            auth=HTTPBasicAuth(config("JIRA_EMAIL"),config("JIRA_API_TOKEN"))
        )
        response.raise_for_status()
    except requests.RequestException as e:
        return None, f"Error al obtener issues: {e}"

    data = response.json()
    detailed_issues = [
        {
            'key': issue['key'],
            'summary': issue['fields'].get('summary'),
            'description': issue['fields'].get('description', 'Sin descripción'),
            'priority': issue['fields']['priority']['name'] if issue['fields'].get('priority') else 'Sin prioridad',
        }
        for issue in data.get('issues', [])
    ]

    return detailed_issues, None